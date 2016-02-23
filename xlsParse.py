#test push


from openpyxl import Workbook
from openpyxl import load_workbook
from loginTest import checkPolycomHDX
import datetime
import os, sys

##define xls path
wb = load_workbook(filename="../source/codecList.xlsx")
tab = wb["Sheet1"]

##check number of rows to review in Excel file. No blank row in between expected
totalRowTab = tab.max_row + 1
print "Number of endpoints to check : ", totalRowTab - 1

rep = "../source/" + str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")) + "/"
lineNumber = 20

os.mkdir(rep, 0755)

#start at row 2 (first row contains generic information)
i = 2

while i < totalRowTab:
    #column "B" contains IP address to use
    cellIP = "B" + str(i)
    #column "C" contains IP address to use
    cellLogin = "C" + str(i)
    #column "D" contains IP address to use
    cellPassword = "D" + str(i)
    #column "E" will be updated with the accessibility of the IP
    cellIPAccess = "E" + str(i)
    #column "F" will be updated with the auth credentials process successfull
    cellWebAccess ="F" + str(i)
    #column "G" will be updated with model
    cellModel = "G" + str(i)
    #column "H" will be updated with SN
    cellSN ="H" + str(i)
    #column "I" will be updated with the timing
    cellTiming = "I" + str(i)
    currentIP = tab[cellIP].value
    currentLogin = tab[cellLogin].value
    currentPassword = tab[cellPassword].value
    #testing each IP ; updating file
    ipCheck = checkPolycomHDX(IP = currentIP, loginSan = currentLogin, passwordSan = currentPassword )
    print ipCheck
    tab[cellIPAccess] = ipCheck["ip"]
    tab[cellWebAccess] = ipCheck["https"]
    tab[cellModel] = ipCheck["model"]
    tab[cellSN] = ipCheck["SN"]
    tab[cellTiming] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if i % lineNumber == 0:
        fileToSave = rep + "endpointResult-" + str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")) + ".xlsx"
        wb.save(fileToSave)
    i = i + 1



fileToSave = rep + "FinalEndpointResult-" + str(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")) + ".xlsx"

#wb.save("source/codecResult.xlsx")
wb.save(fileToSave)